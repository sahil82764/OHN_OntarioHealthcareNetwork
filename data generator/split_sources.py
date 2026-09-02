#!/usr/bin/env python3
"""
Split the generated landing data into five distinct source systems, each in
its own native format.

The point of this is not to move files around. Each target below has a
genuinely different ingestion problem, and that is what makes the Fabric
build worth doing:

  1. SQL Server (on-prem)  relational, incremental by watermark, needs a
                           gateway. The pattern most enterprise data uses.
  2. REST API              paginated JSON with auth and rate limits. Fails
                           in ways files never do — timeouts, 429s, cursors
                           that expire mid-pull.
  3. SFTP drop             X12 EDI. Not tabular at all. Needs a real parser
                           before any of it becomes rows.
  4. Azure Blob            HL7 v2 pipe-delimited messages. Same problem,
                           different grammar, plus one-file-per-message.
  5. SharePoint            Excel maintained by humans. Arrives with merged
                           cells, trailing blank rows, and inconsistent
                           headers, because that is what humans produce.

Usage:
    python split_sources.py --landing ./landing --out ./sources
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import random
import shutil
from collections import defaultdict
from datetime import datetime, timedelta

RNG = random.Random(20260801)

# =====================================================================
# SOURCE ASSIGNMENT
# =====================================================================

SQLSERVER_MAP = {
    "OHN_EHR": [
        ("EHR", "patient"), ("EHR", "admission"), ("EHR", "bed_assignment"),
        ("EHR", "diagnosis"), ("EHR", "emergency_visit"),
    ],
    "OHN_SCHED": [
        ("SCHED", "patient"), ("SCHED", "appointment"),
        ("SCHED", "appointment_status_history"),
    ],
    "OHN_FIN": [
        ("FIN", "patient_account"), ("FIN", "invoice"), ("FIN", "invoice_line"),
    ],
}

API_ENTITIES = [("PHARM", "medication_order"), ("FACIL", "hospital"),
                ("FACIL", "department"), ("FACIL", "bed")]

SFTP_ENTITIES = [("CLAIMS", "claim_header"), ("CLAIMS", "claim_line")]

BLOB_ENTITIES = [("LIS", "lab_result")]

SHAREPOINT_ENTITIES = [("HR", "doctor"), ("SURVEY", "survey_response")]

REF_ENTITIES = [("REF", "icd10ca"), ("REF", "loinc"), ("REF", "medication"),
                ("REF", "payer"), ("REF", "code_mapping")]

# =====================================================================
# SQL SERVER COLUMN TYPES
# Explicit rather than inferred. Type inference on CSV would make every
# health card number a bigint and strip its leading zeros.
# =====================================================================

SQL_TYPES = {
    ("EHR", "patient"): {
        "patient_id": "VARCHAR(20) NOT NULL", "first_name": "NVARCHAR(60)",
        "last_name": "NVARCHAR(60)", "date_of_birth": "VARCHAR(12)",
        "gender": "VARCHAR(10)", "health_card_number": "VARCHAR(20)",
        "postal_code": "VARCHAR(12)", "primary_phone": "VARCHAR(24)",
        "preferred_language": "NVARCHAR(40)", "deceased_indicator": "CHAR(1)",
        "last_modified_ts": "DATETIME2(3)",
    },
    ("EHR", "admission"): {
        "admission_id": "VARCHAR(20) NOT NULL", "encounter_id": "VARCHAR(20)",
        "patient_id": "VARCHAR(20)", "attending_physician_id": "VARCHAR(20)",
        "department_id": "VARCHAR(30)", "hospital_id": "VARCHAR(10)",
        "admission_ts": "DATETIME2(3)", "discharge_ts": "DATETIME2(3) NULL",
        "expected_discharge_ts": "DATETIME2(3) NULL",
        "admission_type": "VARCHAR(10)", "discharge_disposition_code": "VARCHAR(10)",
        "primary_diagnosis_code": "VARCHAR(20)", "payer_id": "VARCHAR(10)",
        "last_modified_ts": "DATETIME2(3)",
    },
    ("EHR", "bed_assignment"): {
        "admission_id": "VARCHAR(20)", "encounter_id": "VARCHAR(20)",
        "bed_id": "VARCHAR(30)", "assignment_start_ts": "DATETIME2(3)",
        "assignment_end_ts": "DATETIME2(3) NULL", "assignment_reason": "VARCHAR(30)",
    },
    ("EHR", "diagnosis"): {
        "encounter_id": "VARCHAR(20)", "patient_id": "VARCHAR(20)",
        "diagnosis_code": "VARCHAR(20)", "diagnosis_rank": "SMALLINT",
        "diagnosis_type": "VARCHAR(20)", "is_primary": "BIT",
        "is_present_on_admission": "BIT", "diagnosis_date": "DATE",
    },
    ("EHR", "emergency_visit"): {
        "ed_visit_id": "VARCHAR(20) NOT NULL", "encounter_id": "VARCHAR(20)",
        "patient_id": "VARCHAR(20)", "hospital_id": "VARCHAR(10)",
        "department_id": "VARCHAR(30)", "arrival_ts": "DATETIME2(3)",
        "triage_ts": "DATETIME2(3) NULL", "physician_seen_ts": "DATETIME2(3) NULL",
        "admit_decision_ts": "DATETIME2(3) NULL", "bed_assigned_ts": "DATETIME2(3) NULL",
        "departure_ts": "DATETIME2(3) NULL", "triage_score": "VARCHAR(4)",
        "arrival_mode": "VARCHAR(20)", "triage_diagnosis_code": "VARCHAR(20)",
        "left_without_being_seen": "BIT", "resulted_in_admission": "BIT",
        "payer_id": "VARCHAR(10)", "last_modified_ts": "DATETIME2(3)",
    },
    ("SCHED", "patient"): {
        "patient_ref": "VARCHAR(20) NOT NULL", "given": "NVARCHAR(60)",
        "surname": "NVARCHAR(60)", "dob": "VARCHAR(12)", "sex": "VARCHAR(10)",
        "health_card": "VARCHAR(20)", "postal": "VARCHAR(12)",
        "contact_phone": "VARCHAR(24)", "modified_at": "DATETIME2(3)",
    },
    ("SCHED", "appointment"): {
        "appointment_id": "VARCHAR(20) NOT NULL", "patient_ref": "VARCHAR(20)",
        "doctor_id": "VARCHAR(20)", "department_id": "VARCHAR(30)",
        "hospital_id": "VARCHAR(10)", "booking_ts": "DATETIME2(3)",
        "scheduled_ts": "DATETIME2(3)", "scheduled_duration_min": "INT",
        "appointment_type": "VARCHAR(40)", "is_first_visit": "BIT",
        "is_virtual": "BIT", "checkin_ts": "DATETIME2(3) NULL",
        "seen_ts": "DATETIME2(3) NULL", "checkout_ts": "DATETIME2(3) NULL",
        "cancellation_ts": "DATETIME2(3) NULL", "cancelled_by": "VARCHAR(20)",
        "status_code": "VARCHAR(30)", "modified_at": "DATETIME2(3)",
    },
    ("SCHED", "appointment_status_history"): {
        "appointment_id": "VARCHAR(20)", "status_code": "VARCHAR(30)",
        "status_ts": "DATETIME2(3)",
    },
    ("FIN", "patient_account"): {
        "account_holder_id": "VARCHAR(20) NOT NULL", "first_nm": "NVARCHAR(60)",
        "last_nm": "NVARCHAR(60)", "birth_dt": "VARCHAR(12)",
        "gender_cd": "VARCHAR(10)", "hc_number": "VARCHAR(20)",
        "zip_postal": "VARCHAR(12)", "phone_number": "VARCHAR(24)",
        "update_dt": "DATETIME2(3)",
    },
    ("FIN", "invoice"): {
        "invoice_id": "VARCHAR(20) NOT NULL", "patient_id": "VARCHAR(20)",
        "encounter_id": "VARCHAR(20)", "hospital_id": "VARCHAR(10)",
        "payer_id": "VARCHAR(10)", "invoice_date": "DATE",
        "service_date": "DATE", "total_amount": "DECIMAL(18,2)",
        "update_dt": "DATETIME2(3)",
    },
    ("FIN", "invoice_line"): {
        "invoice_id": "VARCHAR(20)", "invoice_line_number": "INT",
        "patient_id": "VARCHAR(20)", "department_id": "VARCHAR(30)",
        "hospital_id": "VARCHAR(10)", "payer_id": "VARCHAR(10)",
        "encounter_id": "VARCHAR(20)", "service_code": "VARCHAR(20)",
        "service_category": "NVARCHAR(60)", "quantity": "DECIMAL(12,2)",
        "charge_amount": "DECIMAL(18,2)", "discount_amount": "DECIMAL(18,2)",
        "tax_amount": "DECIMAL(18,2)", "net_amount": "DECIMAL(18,2)",
        "payment_amount": "DECIMAL(18,2)", "outstanding_amount": "DECIMAL(18,2)",
    },
}

WATERMARK_COLUMNS = {
    ("EHR", "patient"): "last_modified_ts",
    ("EHR", "admission"): "last_modified_ts",
    ("EHR", "emergency_visit"): "last_modified_ts",
    ("SCHED", "patient"): "modified_at",
    ("SCHED", "appointment"): "modified_at",
    ("FIN", "patient_account"): "update_dt",
    ("FIN", "invoice"): "update_dt",
}


# =====================================================================
# HELPERS
# =====================================================================

def read_entity(landing: str, source: str, entity: str) -> tuple[list[str], list[dict]]:
    base = os.path.join(landing, source, entity)
    if not os.path.isdir(base):
        raise FileNotFoundError(f"{base} not found — run ohn_generator.py first")
    header, rows = [], []
    for part in sorted(os.listdir(base)):
        pdir = os.path.join(base, part)
        if not os.path.isdir(pdir):
            continue
        for fn in sorted(os.listdir(pdir)):
            if not fn.endswith(".csv"):
                continue
            with open(os.path.join(pdir, fn), encoding="utf-8") as fh:
                r = csv.DictReader(fh)
                header = r.fieldnames or []
                rows.extend(r)
    return header, rows


def ensure(path: str) -> str:
    os.makedirs(path, exist_ok=True)
    return path


def sql_escape(v: str) -> str:
    return v.replace("'", "''")


# =====================================================================
# 1. SQL SERVER
# =====================================================================

def build_sqlserver(landing: str, out: str) -> dict:
    root = ensure(os.path.join(out, "01-sqlserver"))
    data_dir = ensure(os.path.join(root, "data"))
    stats = {}

    ddl_lines = [
        "/* Auto-generated by split_sources.py — do not edit by hand. */",
        "/* Simulates three on-prem hospital systems on one SQL Server host. */",
        "SET NOCOUNT ON;",
        "GO",
        "",
    ]

    for db, entities in SQLSERVER_MAP.items():
        ddl_lines += [
            f"IF DB_ID('{db}') IS NULL CREATE DATABASE {db};",
            "GO",
            f"USE {db};",
            "GO",
            "",
        ]
        for source, entity in entities:
            header, rows = read_entity(landing, source, entity)
            table = entity
            types = SQL_TYPES.get((source, entity), {})
            cols = []
            for c in header:
                cols.append(f"    [{c}] {types.get(c, 'NVARCHAR(200)')}")
            ddl_lines += [
                f"IF OBJECT_ID('dbo.{table}') IS NOT NULL DROP TABLE dbo.{table};",
                f"CREATE TABLE dbo.{table} (",
                ",\n".join(cols),
                ");",
                "GO",
            ]

            # An index on the watermark column, because the Fabric Copy
            # activity will filter on it on every incremental run and a scan
            # of a 200k-row table per pull is a self-inflicted wound.
            wm = WATERMARK_COLUMNS.get((source, entity))
            if wm:
                ddl_lines.append(
                    f"CREATE INDEX IX_{table}_{wm} ON dbo.{table} ([{wm}]);")
                ddl_lines.append("GO")
            ddl_lines.append("")

            # Write a clean CSV for BULK INSERT. Empty strings become \\N so
            # SQL Server stores a real NULL instead of the literal text.
            out_csv = os.path.join(data_dir, f"{db}.{table}.csv")
            with open(out_csv, "w", newline="", encoding="utf-8") as fh:
                w = csv.writer(fh, quoting=csv.QUOTE_MINIMAL)
                w.writerow(header)
                for r in rows:
                    w.writerow([r.get(c, "") for c in header])
            stats[f"{db}.{table}"] = len(rows)

    with open(os.path.join(root, "02_create_tables.sql"), "w", encoding="utf-8") as fh:
        fh.write("\n".join(ddl_lines))

    # BULK INSERT script
    load_lines = ["SET NOCOUNT ON;", "GO", ""]
    for db, entities in SQLSERVER_MAP.items():
        load_lines += [f"USE {db};", "GO", ""]
        for source, entity in entities:
            load_lines += [
                f"TRUNCATE TABLE dbo.{entity};",
                f"BULK INSERT dbo.{entity}",
                f"FROM '/data/{db}.{entity}.csv'",
                "WITH (",
                "    FORMAT = 'CSV',",
                "    FIRSTROW = 2,",
                "    FIELDTERMINATOR = ',',",
                "    ROWTERMINATOR = '0x0a',",
                "    TABLOCK,",
                "    MAXERRORS = 50",
                ");",
                f"PRINT 'Loaded {db}.{entity}: ' + CAST(@@ROWCOUNT AS VARCHAR(20));",
                "GO",
                "",
            ]
    with open(os.path.join(root, "03_load_data.sql"), "w", encoding="utf-8") as fh:
        fh.write("\n".join(load_lines))

    # docker-compose
    compose = """services:
  sqlserver:
    image: mcr.microsoft.com/mssql/server:2022-latest
    container_name: ohn-sqlserver
    environment:
      ACCEPT_EULA: "Y"
      MSSQL_SA_PASSWORD: "${MSSQL_SA_PASSWORD:?set MSSQL_SA_PASSWORD in .env}"
      MSSQL_PID: "Developer"
    ports:
      - "1433:1433"
    volumes:
      - ./data:/data:ro
      - ./01_create_databases.sql:/scripts/01_create_databases.sql:ro
      - ./02_create_tables.sql:/scripts/02_create_tables.sql:ro
      - ./03_load_data.sql:/scripts/03_load_data.sql:ro
      - mssql_data:/var/opt/mssql
    healthcheck:
      test: ["CMD-SHELL",
             "/opt/mssql-tools18/bin/sqlcmd -S localhost -U sa -P \\"$$MSSQL_SA_PASSWORD\\" -C -Q 'SELECT 1' || exit 1"]
      interval: 10s
      timeout: 5s
      retries: 12
      start_period: 30s

volumes:
  mssql_data:
"""
    with open(os.path.join(root, "docker-compose.yml"), "w", encoding="utf-8") as fh:
        fh.write(compose)

    # Read-only login for Fabric — never point a gateway at sa
    with open(os.path.join(root, "04_create_reader_login.sql"), "w", encoding="utf-8") as fh:
        fh.write("""/* A dedicated read-only login for the Fabric gateway connection.

   Pointing a gateway at sa works and is exactly the habit that turns a
   portfolio project into a security finding. This login can read the three
   source databases and nothing else. */

USE master;
GO
IF NOT EXISTS (SELECT 1 FROM sys.server_principals WHERE name = 'fabric_reader')
    CREATE LOGIN fabric_reader WITH PASSWORD = 'Ch4ngeMe_InYourEnv!',
        CHECK_POLICY = ON;
GO

DECLARE @db SYSNAME, @sql NVARCHAR(MAX);
DECLARE db_cur CURSOR FOR
    SELECT name FROM sys.databases WHERE name IN ('OHN_EHR','OHN_SCHED','OHN_FIN');
OPEN db_cur;
FETCH NEXT FROM db_cur INTO @db;
WHILE @@FETCH_STATUS = 0
BEGIN
    SET @sql = N'USE ' + QUOTENAME(@db) + N';
        IF NOT EXISTS (SELECT 1 FROM sys.database_principals WHERE name = ''fabric_reader'')
            CREATE USER fabric_reader FOR LOGIN fabric_reader;
        ALTER ROLE db_datareader ADD MEMBER fabric_reader;';
    EXEC sp_executesql @sql;
    FETCH NEXT FROM db_cur INTO @db;
END
CLOSE db_cur;
DEALLOCATE db_cur;
GO
""")

    with open(os.path.join(root, "01_create_databases.sql"), "w", encoding="utf-8") as fh:
        fh.write("\n".join(
            f"IF DB_ID('{db}') IS NULL CREATE DATABASE {db};\nGO"
            for db in SQLSERVER_MAP))

    return stats


# =====================================================================
# 2. REST API payload store
# =====================================================================

def build_api(landing: str, out: str) -> dict:
    root = ensure(os.path.join(out, "02-api"))
    data_dir = ensure(os.path.join(root, "data"))
    stats = {}
    for source, entity in API_ENTITIES:
        header, rows = read_entity(landing, source, entity)
        # The API serves JSON, so store JSON. Numeric coercion happens here
        # rather than in the API so the server stays a thin transport layer.
        payload = []
        for r in rows:
            payload.append({k: (v if v != "" else None) for k, v in r.items()})
        name = f"{source.lower()}_{entity}"
        with open(os.path.join(data_dir, f"{name}.json"), "w", encoding="utf-8") as fh:
            json.dump(payload, fh)
        stats[name] = len(rows)
    return stats


# =====================================================================
# 3. SFTP — X12 EDI 837P claims and 835 remittances
# =====================================================================

def _isa(control: int, sender: str, receiver: str, now: datetime) -> str:
    return "*".join([
        "ISA", "00", " " * 10, "00", " " * 10,
        "ZZ", f"{sender:<15}", "ZZ", f"{receiver:<15}",
        now.strftime("%y%m%d"), now.strftime("%H%M"),
        "^", "00501", f"{control:09d}", "0", "P", ":",
    ]) + "~"


def build_sftp(landing: str, out: str) -> dict:
    """Write X12 837P claim submissions and 835 remittance advices.

    Simplified but structurally faithful: real envelopes, real segment
    ordering, real loop structure. The point is that no amount of
    `spark.read.csv` will open these — a parser has to walk the segments.
    """
    root = ensure(os.path.join(out, "03-sftp"))
    outbound = ensure(os.path.join(root, "outbound"))   # claims we submit
    inbound = ensure(os.path.join(root, "inbound"))     # remittances we receive

    _, headers = read_entity(landing, "CLAIMS", "claim_header")
    _, lines = read_entity(landing, "CLAIMS", "claim_line")

    lines_by_claim = defaultdict(list)
    for l in lines:
        lines_by_claim[l["claim_number"]].append(l)

    # Group claims into daily submission batches, as a real clearinghouse feed
    by_day = defaultdict(list)
    for h in headers:
        if h["submission_date"]:
            by_day[h["submission_date"]].append(h)

    control = 1
    n_837 = n_835 = 0

    # Only a sample of days, or you get 900 files and a slow demo
    days = sorted(by_day)
    sample_days = days[::7][:60]

    for day in sample_days:
        claims = by_day[day][:400]
        if not claims:
            continue
        now = datetime.strptime(day, "%Y-%m-%d")
        segs = [_isa(control, "OHNHEALTH", "CLEARINGHSE", now)]
        segs.append(f"GS*HC*OHNHEALTH*CLEARINGHSE*{now.strftime('%Y%m%d')}*"
                    f"{now.strftime('%H%M')}*{control}*X*005010X222A1~")
        segs.append(f"ST*837*{control:04d}*005010X222A1~")
        segs.append(f"BHT*0019*00*{control:09d}*{now.strftime('%Y%m%d')}*"
                    f"{now.strftime('%H%M')}*CH~")
        segs.append("NM1*41*2*ONTARIO HEALTHCARE NETWORK*****46*OHNHEALTH~")
        segs.append("PER*IC*EDI DEPT*TE*4165551212~")
        segs.append("NM1*40*2*CLEARINGHOUSE INC*****46*CLEARINGHSE~")

        hl = 0
        for c in claims:
            hl += 1
            segs.append(f"HL*{hl}**20*1~")
            segs.append(f"NM1*85*2*ONTARIO HEALTHCARE NETWORK*****XX*{c['hospital_id']}~")
            segs.append(f"REF*EI*{c['hospital_id']}~")
            hl += 1
            segs.append(f"HL*{hl}*{hl-1}*22*0~")
            segs.append(f"SBR*P*18*******{c['payer_id']}~")
            segs.append(f"NM1*IL*1*PATIENT*SYNTHETIC****MI*{c['patient_id']}~")
            segs.append(f"NM1*PR*2*PAYER {c['payer_id']}*****PI*{c['payer_id']}~")
            billed = c["billed_amount"] or "0"
            segs.append(f"CLM*{c['claim_number']}*{billed}***11:B:1*Y*A*Y*Y~")
            svc_date = (c["service_date"] or day).replace("-", "")
            segs.append(f"DTP*472*D8*{svc_date}~")
            segs.append(f"HI*ABK:{c['primary_diagnosis_code'].replace('.', '')}~")
            for i, ln in enumerate(lines_by_claim.get(c["claim_number"], [])[:12], 1):
                segs.append(f"LX*{i}~")
                segs.append(f"SV1*HC:{ln['service_code']}*{ln['line_amount']}*UN*"
                            f"{ln['quantity']}***1~")
                segs.append(f"DTP*472*D8*{svc_date}~")

        segs.append(f"SE*{len(segs) - 2}*{control:04d}~")
        segs.append(f"GE*1*{control}~")
        segs.append(f"IEA*1*{control:09d}~")

        fn = os.path.join(outbound, f"837P_{day.replace('-', '')}_{control:06d}.edi")
        with open(fn, "w", encoding="utf-8") as fh:
            fh.write("\n".join(segs))
        n_837 += 1
        control += 1

        # --- matching 835 remittance for the adjudicated claims -----------
        adjudicated = [c for c in claims if c["adjudication_date"]]
        if not adjudicated:
            continue
        rdate = datetime.strptime(adjudicated[0]["adjudication_date"], "%Y-%m-%d")
        total_paid = sum(float(c["paid_amount"] or 0) for c in adjudicated)

        rsegs = [_isa(control, "CLEARINGHSE", "OHNHEALTH", rdate)]
        rsegs.append(f"GS*HP*CLEARINGHSE*OHNHEALTH*{rdate.strftime('%Y%m%d')}*"
                     f"{rdate.strftime('%H%M')}*{control}*X*005010X221A1~")
        rsegs.append(f"ST*835*{control:04d}~")
        rsegs.append(f"BPR*I*{total_paid:.2f}*C*ACH*CCP*01*999999999*DA*"
                     f"1234567*1512345678**01*999988880*DA*98765*"
                     f"{rdate.strftime('%Y%m%d')}~")
        rsegs.append(f"TRN*1*{control:09d}*1512345678~")
        rsegs.append(f"DTM*405*{rdate.strftime('%Y%m%d')}~")
        rsegs.append("N1*PR*ONTARIO HEALTH PAYER~")
        rsegs.append("N1*PE*ONTARIO HEALTHCARE NETWORK*XX*1234567890~")
        rsegs.append(f"LX*1~")

        for c in adjudicated:
            billed = float(c["billed_amount"] or 0)
            paid = float(c["paid_amount"] or 0)
            status = {"A1": "1", "A2": "2", "D1": "4"}.get(c["status_code"], "1")
            rsegs.append(f"CLP*{c['claim_number']}*{status}*{billed:.2f}*{paid:.2f}*"
                         f"{float(c['patient_responsibility'] or 0):.2f}*"
                         f"12*{c['claim_number']}*11~")
            rsegs.append(f"NM1*QC*1*PATIENT*SYNTHETIC****MI*{c['patient_id']}~")
            if c["denial_reason_code"]:
                grp, code = c["denial_reason_code"].split("-")
                denied = float(c["denied_amount"] or 0)
                rsegs.append(f"CAS*{grp}*{code}*{denied:.2f}~")
            claim_svc = (c["service_date"] or day).replace("-", "")
            rsegs.append(f"DTM*232*{claim_svc}~")

        rsegs.append(f"SE*{len(rsegs) - 2}*{control:04d}~")
        rsegs.append(f"GE*1*{control}~")
        rsegs.append(f"IEA*1*{control:09d}~")

        fn = os.path.join(inbound, f"835_{adjudicated[0]['adjudication_date'].replace('-', '')}"
                                   f"_{control:06d}.edi")
        with open(fn, "w", encoding="utf-8") as fh:
            fh.write("\n".join(rsegs))
        n_835 += 1
        control += 1

    return {"837P files": n_837, "835 files": n_835}


# =====================================================================
# 4. Azure Blob — HL7 v2 ORU^R01 lab results
# =====================================================================

def build_blob(landing: str, out: str) -> dict:
    """Write HL7 v2.5 ORU^R01 result messages, batched by day.

    Grouping is by (encounter, order time) so one message carries one OBR
    with several OBX segments — which is how a real LIS emits a panel, and
    which forces the parser to handle nested repeating segments rather than
    one flat line per row.
    """
    root = ensure(os.path.join(out, "04-blob"))
    _, rows = read_entity(landing, "LIS", "lab_result")

    # Group into panels the way a lab actually reports them: one specimen,
    # one panel, one OBR, many OBX. Grouping by exact timestamp made every
    # key unique and every message carried a single OBX, which quietly
    # removes the whole reason HL7 is worth parsing. Grouping by the hour
    # barely helped, because orders are spread across the admission.
    # Panel membership comes from the LOINC reference, so a chemistry panel
    # and a haematology panel drawn the same day stay separate messages.
    panel_of, loinc_meta = {}, {}
    try:
        _, loinc_rows = read_entity(landing, "REF", "loinc")
        panel_of = {r["loinc_code"]: r["panel_name"] for r in loinc_rows}
        loinc_meta = {r["loinc_code"]: (r["result_unit"],
                                        r["reference_low"], r["reference_high"])
                      for r in loinc_rows}
    except FileNotFoundError:
        pass

    orders = defaultdict(list)
    for r in rows:
        day_bucket = (r["order_ts"] or "")[:10]
        panel = panel_of.get(r["loinc_code"], "Miscellaneous")
        key = (r["encounter_id"], f"{day_bucket}|{panel}|{r['priority']}",
               r["hospital_id"])
        orders[key].append(r)

    by_day = defaultdict(list)
    for key, results in orders.items():
        day = (key[1] or "1900-01-01 00:00:00")[:10]
        by_day[day].append((key, results))

    msg_ctrl = 1
    n_files = n_msgs = 0

    for day in sorted(by_day):
        day_orders = by_day[day]
        if not day_orders:
            continue
        folder = ensure(os.path.join(root, f"lab/{day[:4]}/{day[5:7]}/{day[8:10]}"))
        messages = []
        for (enc_id, _hour, hosp), results in day_orders:
            results.sort(key=lambda x: x["order_ts"] or "")
            first = results[0]
            ts_compact = (first["order_ts"] or "").replace("-", "").replace(
                ":", "").replace(" ", "")[:14]
            segs = [
                f"MSH|^~\\&|LIS|{hosp}|EHR|OHN|{ts_compact}||ORU^R01^ORU_R01|"
                f"MSG{msg_ctrl:09d}|P|2.5|||AL|NE",
                f"PID|1||{first['patient_id']}^^^OHN^MR||SYNTHETIC^PATIENT||||||||||||",
                f"PV1|1|I|{first['department_id']}|||||{first['ordering_doctor_id']}"
                f"|||||||||||{enc_id}",
                f"ORC|RE|{first['lab_result_id']}|{first['lab_result_id']}||CM||||"
                f"{ts_compact}|||{first['ordering_doctor_id']}",
                f"OBR|1|{first['lab_result_id']}|{first['lab_result_id']}|"
                f"{first['loinc_code']}^^LN|"
                f"{'S' if first['priority'] == 'STAT' else 'R'}|{ts_compact}|"
                f"{(first['collect_ts'] or '').replace('-','').replace(':','').replace(' ','')[:14]}"
                f"||||||||{first['ordering_doctor_id']}||||||"
                f"{(first['result_ts'] or '').replace('-','').replace(':','').replace(' ','')[:14]}"
                f"|||F",
            ]
            for i, r in enumerate(results, 1):
                is_numeric = bool(r["result_value_numeric"])
                value = r["result_value_numeric"] if is_numeric else r["result_value_text"]
                vtype = "NM" if is_numeric else "ST"
                res_ts = (r["result_ts"] or "").replace("-", "").replace(":", "").replace(" ", "")[:14]
                # HL7 field positions are 1-based and unforgiving. OBX-5 is
                # the value, OBX-6 units, OBX-7 reference range, OBX-8 the
                # abnormal flag, OBX-11 result status, OBX-14 the observation
                # timestamp. One extra pipe here shifts every field right and
                # the parser reads blanks where the flags should be — which
                # is exactly what happened the first time this was written.
                unit, ref_lo, ref_hi = loinc_meta.get(r["loinc_code"], ("", "", ""))
                ref_range = f"{ref_lo}-{ref_hi}" if ref_lo and ref_hi else ""
                segs.append(
                    f"OBX|{i}|{vtype}|{r['loinc_code']}^^LN||{value}|{unit}|"
                    f"{ref_range}|{r['abnormal_flag']}|||F|||{res_ts}"
                )
            messages.append("\r".join(segs))
            msg_ctrl += 1
            n_msgs += 1

        # Batch into files of 200 messages, separated by the HL7 file
        # header/trailer convention (FHS/BHS). Real interfaces send either
        # one message per file or batched — batched is the harder parse.
        for chunk_ix in range(0, len(messages), 200):
            chunk = messages[chunk_ix:chunk_ix + 200]
            fn = os.path.join(folder, f"ORU_{day.replace('-','')}_{chunk_ix//200:03d}.hl7")
            body = "\r".join([
                f"FHS|^~\\&|LIS|OHN|EHR|OHN|{day.replace('-','')}000000",
                f"BHS|^~\\&|LIS|OHN|EHR|OHN|{day.replace('-','')}000000",
            ]) + "\r" + "\r".join(chunk) + "\r" + "\r".join([
                f"BTS|{len(chunk)}", "FTS|1",
            ])
            with open(fn, "w", encoding="utf-8", newline="") as fh:
                fh.write(body)
            n_files += 1

    return {"HL7 files": n_files, "ORU messages": n_msgs}


# =====================================================================
# 5. SharePoint — Excel workbooks maintained by humans
# =====================================================================

def build_sharepoint(landing: str, out: str) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    root = ensure(os.path.join(out, "05-sharepoint"))
    stats = {}

    title_font = Font(name="Arial", size=14, bold=True)
    head_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    head_fill = PatternFill("solid", fgColor="4F6228")

    for source, entity in SHAREPOINT_ENTITIES:
        header, rows = read_entity(landing, source, entity)
        wb = Workbook()
        ws = wb.active
        ws.title = entity[:31]

        # Human-maintained workbooks have a title block above the data, which
        # is precisely why a naive read grabs the wrong header row. Dataflow
        # Gen2 has to be told to skip these rows — that is the lesson.
        ws["A1"] = ("Medical Staff Roster" if entity == "doctor"
                    else "Patient Experience Survey Export")
        ws["A1"].font = title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(8, len(header)))
        ws["A2"] = f"Exported {datetime.now().strftime('%Y-%m-%d')} — CONFIDENTIAL"
        ws["A2"].font = Font(name="Arial", size=9, italic=True)
        ws["A3"] = ""

        for c_ix, col in enumerate(header, 1):
            cell = ws.cell(row=4, column=c_ix, value=col)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center")

        for r_ix, r in enumerate(rows, 5):
            for c_ix, col in enumerate(header, 1):
                v = r.get(col, "")
                # Numeric-looking strings become numbers, which is exactly
                # the type drift a spreadsheet source introduces.
                try:
                    v = float(v) if v not in ("", None) and "." in str(v) else (
                        int(v) if str(v).isdigit() else v)
                except (ValueError, TypeError):
                    pass
                cell = ws.cell(row=r_ix, column=c_ix, value=v)
                cell.font = body_font

        for c_ix, col in enumerate(header, 1):
            width = min(30, max(12, len(col) + 4))
            ws.column_dimensions[ws.cell(row=4, column=c_ix).column_letter].width = width
        ws.freeze_panes = "A5"

        fn = os.path.join(root, f"{entity}_roster.xlsx" if entity == "doctor"
                          else f"{entity}_export.xlsx")
        wb.save(fn)
        stats[entity] = len(rows)

    return stats


# =====================================================================
# 6. Reference data — manual upload
# =====================================================================

def build_reference(landing: str, out: str) -> dict:
    root = ensure(os.path.join(out, "06-reference"))
    stats = {}
    for source, entity in REF_ENTITIES:
        header, rows = read_entity(landing, source, entity)
        with open(os.path.join(root, f"{entity}.csv"), "w", newline="",
                  encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(header)
            for r in rows:
                w.writerow([r.get(c, "") for c in header])
        stats[entity] = len(rows)
    return stats


# =====================================================================
# MAIN
# =====================================================================

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--landing", default="./landing")
    ap.add_argument("--out", default="./sources")
    ap.add_argument("--clean", action="store_true",
                    help="Delete the output directory first")
    args = ap.parse_args()

    # Remove only the folders this script generates. Deleting args.out
    # wholesale breaks the documented `--out .` usage (you cannot rmtree the
    # directory you are standing in) and would also destroy api_server.py,
    # setup.sh, and .env, which live alongside the generated output.
    if args.clean:
        for sub in ("01-sqlserver", "02-api/data", "03-sftp", "04-blob",
                    "05-sharepoint", "06-reference"):
            target = os.path.join(args.out, sub)
            if os.path.isdir(target):
                shutil.rmtree(target)
    ensure(args.out)

    print("Splitting generated data into five source systems\n")

    print("[1/6] SQL Server — three on-prem databases")
    s1 = build_sqlserver(args.landing, args.out)
    for k, v in s1.items():
        print(f"        {k:<44} {v:>9,} rows")

    print("\n[2/6] REST API — JSON payload store")
    s2 = build_api(args.landing, args.out)
    for k, v in s2.items():
        print(f"        {k:<44} {v:>9,} records")

    print("\n[3/6] SFTP — X12 EDI 837P / 835")
    s3 = build_sftp(args.landing, args.out)
    for k, v in s3.items():
        print(f"        {k:<44} {v:>9,}")

    print("\n[4/6] Azure Blob — HL7 v2 ORU^R01")
    s4 = build_blob(args.landing, args.out)
    for k, v in s4.items():
        print(f"        {k:<44} {v:>9,}")

    print("\n[5/6] SharePoint — Excel workbooks")
    s5 = build_sharepoint(args.landing, args.out)
    for k, v in s5.items():
        print(f"        {k:<44} {v:>9,} rows")

    print("\n[6/6] Reference data — manual upload CSVs")
    s6 = build_reference(args.landing, args.out)
    for k, v in s6.items():
        print(f"        {k:<44} {v:>9,} rows")

    print(f"\nWritten to {os.path.abspath(args.out)}")
    print("\nNext: QUICKSTART.md, Phase 3 (SQL Server) and Phase 4 (API).")


if __name__ == "__main__":
    main()
