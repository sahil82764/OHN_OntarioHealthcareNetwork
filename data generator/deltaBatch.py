#!/usr/bin/env python3
"""
Ontario Healthcare Network — delta batch generator.

Produces an INCREMENTAL batch on top of an existing generated dataset, then
distributes it to the five source systems.

Why a delta batch rather than a second full generation
------------------------------------------------------
A fresh load re-tests nothing. This batch exercises the parts of the platform
that only run on the second pass:

  * watermarks              incremental entities should pick up only new rows
  * SCD2 versioning         changed attributes must create a second dimension
                            row, not overwrite the first
  * MPI stability           golden IDs must survive reprocessing; new records
                            for known people must join their existing cluster
  * cluster merging         two previously-separate clusters that now link
                            should retire one golden ID and record it
  * deduplication           a re-delivered row must not double the count
  * unknown members         a code that is not in the reference data must
                            resolve to -1, not fail the load

New inconsistencies are introduced that the first batch did not contain, so
the run finds problems rather than confirming old ones.

Usage:
    python generate_delta_batch.py --landing ./landing --out ./landing-delta
    python generate_delta_batch.py --landing ./landing --out ./landing-delta \\
           --apply ../sources          # also update the five source systems
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import random
import shutil
from collections import defaultdict
from datetime import date, datetime, timedelta

import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from ohn_generator import (  # noqa: E402
    DIAGNOSES, FSA_POOL, GIVEN_F, GIVEN_M, FAMILY, LAB_TESTS, MEDICATIONS,
    NICKNAMES, PAYERS, business_hour, diurnal_hour, fmt_d, fmt_ts,
    lognormal_days, make_hcn, make_phone, make_postal, typo,
)

SEED = 20260901

# =====================================================================
# DELTA-SPECIFIC INCONSISTENCIES
#
# These are deliberately different from the first batch. The point is to
# find new problems, not re-confirm known ones.
# =====================================================================

DELTA_DEFECTS = {
    # --- identity churn: the MPI's hardest new cases -----------------
    "surname_change":        0.030,  # marriage/divorce — same person, new family name
    "cross_source_new":      0.120,  # existing patient appears in a system they were absent from
    "new_intra_dup":         0.035,  # existing patient re-registered in the EHR under a new MRN
    "cluster_merge":         0.010,  # a record that links two previously-separate clusters

    # --- attribute drift: should trigger SCD2 versions ---------------
    "address_change":        0.080,  # patient moved — new postal code
    "language_change":       0.015,
    "doctor_dept_move":      0.060,  # doctor changed department
    "doctor_fte_change":     0.040,
    "bed_ward_change":       0.050,  # bed reassigned to a different ward

    # --- reference gaps: should land on the unknown member -----------
    "unknown_diagnosis":     0.020,  # ICD code absent from ref_icd10ca
    "unknown_loinc":         0.018,
    "unknown_din":           0.015,
    "new_sex_code":          0.012,  # a code not in ref_code_mapping

    # --- format and encoding problems the first batch lacked ---------
    "trailing_whitespace":   0.040,  # "  SMITH  " — breaks naive equality joins
    "unicode_lookalike":     0.008,  # Cyrillic А for Latin A — visually identical
    "mixed_case_id":         0.020,  # mrn74864960 vs MRN74864960
    "tz_shifted_ts":         0.015,  # timestamp written in local rather than UTC
    "thousand_separator":    0.010,  # "1,250.00" in a numeric field

    # --- temporal problems -------------------------------------------
    "late_arriving":         0.060,  # backdated event, older than the watermark
    "future_dated":          0.004,  # event dated after today
    "duplicate_delivery":    0.025,  # exact row re-sent, should dedupe on _row_hash
}

# Codes that are deliberately NOT in the reference data
ORPHAN_DIAGNOSES = ["Z99.9", "U09.9", "X99.1", "Q88.8", "R99"]
ORPHAN_LOINC = ["99999-9", "88888-8", "LOCAL-777"]
ORPHAN_DIN = ["09999999", "08888888"]
NEW_SEX_CODES = ["NB", "PNTS", "9", "O"]


def rand_ts(rng, start: date, end: date) -> datetime:
    span = (end - start).days
    d = start + timedelta(days=rng.randrange(max(1, span)))
    return datetime(d.year, d.month, d.day,
                    rng.randrange(24), rng.randrange(60), rng.randrange(60))


def corrupt_text(rng, value: str) -> str:
    """Apply a delta-batch text defect. Returns the value unchanged most of
    the time; the caller decides how often to invoke it."""
    r = rng.random()
    if r < 0.45:
        return f"  {value}  "                      # trailing whitespace
    if r < 0.70:
        return value.lower()                        # case drift
    # Cyrillic lookalikes — visually identical, different code points. These
    # break equality joins in a way that is genuinely hard to spot by eye.
    return (value.replace("A", "\u0410", 1).replace("E", "\u0415", 1)
                 .replace("O", "\u041E", 1))


class DeltaBatch:

    def __init__(self, landing: str, out: str, days: int, seed: int = SEED):
        self.rng = random.Random(seed)
        self.landing = landing
        self.out = out
        self.days = days
        self.tables: dict[str, tuple[list[str], list[list]]] = {}
        self.stats: dict[str, int] = defaultdict(int)

        # The delta window starts where the original data ended.
        self.window_end = date.today()
        self.window_start = self.window_end - timedelta(days=days)

    # ---------------------------------------------------------- helpers
    def read(self, source: str, entity: str) -> tuple[list[str], list[dict]]:
        base = os.path.join(self.landing, source, entity)
        if not os.path.isdir(base):
            return [], []
        header, rows = [], []
        for part in sorted(os.listdir(base)):
            pdir = os.path.join(base, part)
            if not os.path.isdir(pdir):
                continue
            for fn in sorted(os.listdir(pdir)):
                if fn.endswith(".csv"):
                    with open(os.path.join(pdir, fn), encoding="utf-8") as fh:
                        r = csv.DictReader(fh)
                        header = r.fieldnames or []
                        rows.extend(r)
        return header, rows

    def add(self, name: str, header: list[str], rows: list[list]):
        self.tables[name] = (header, rows)

    def chance(self, key: str) -> bool:
        return self.rng.random() < DELTA_DEFECTS[key]

    def event_ts(self) -> str:
        """A timestamp inside the delta window, with occasional temporal
        defects. Late-arriving events are backdated before the watermark —
        the pipeline must still pick them up on a later run."""
        if self.chance("late_arriving"):
            d = self.window_start - timedelta(days=self.rng.randrange(30, 400))
            self.stats["late_arriving"] += 1
        elif self.chance("future_dated"):
            d = self.window_end + timedelta(days=self.rng.randrange(1, 60))
            self.stats["future_dated"] += 1
        else:
            d = self.window_start + timedelta(
                days=self.rng.randrange(max(1, self.days)))
        ts = datetime(d.year, d.month, d.day,
                      self.rng.randrange(24), self.rng.randrange(60),
                      self.rng.randrange(60))
        if self.chance("tz_shifted_ts"):
            # Written in local time rather than UTC. Silver cannot detect
            # this; it shows up as ED waits that are five hours long.
            ts += timedelta(hours=self.rng.choice([-5, -4, 5, 8]))
            self.stats["tz_shifted"] += 1
        return fmt_ts(ts)

    # ================================================== PATIENTS
    def gen_patients(self):
        """Three kinds of patient record in the delta:

        1. Genuinely new people
        2. Updated attributes for existing people — should create SCD2 versions
        3. Existing people appearing in a system they were absent from, or
           re-registered under a new id — the MPI must fold these into the
           existing cluster rather than creating a new golden ID
        """
        rng = self.rng
        _, ehr = self.read("EHR", "patient")
        _, sched = self.read("SCHED", "patient")
        _, fin = self.read("FIN", "patient_account")

        ehr_by_id = {r["patient_id"]: r for r in ehr}
        sched_ids = {r["patient_ref"] for r in sched}
        fin_ids = {r["account_holder_id"] for r in fin}

        # Match a scheduling/finance record back to its EHR record by health
        # card, so cross-source additions reference the right person.
        hcn_to_ehr = {}
        for r in ehr:
            h = (r.get("health_card_number") or "").replace(" ", "").replace("-", "")
            if h:
                hcn_to_ehr[h] = r

        new_ehr, new_sched, new_fin = [], [], []

        # ---- 1. brand-new patients -----------------------------------
        n_new = max(50, int(len(ehr) * 0.04))
        for i in range(n_new):
            sex = rng.choices(["M", "F", "X"], weights=[48.5, 51, 0.5])[0]
            given = rng.choice(GIVEN_M if sex == "M" else GIVEN_F)
            family = rng.choice(FAMILY)
            age = int(min(99, max(0, rng.gauss(52, 22))))
            birth = date.today() - timedelta(days=age * 365 + rng.randrange(365))
            fsa = rng.choice(FSA_POOL)
            hcn = make_hcn(rng)
            mrn = f"MRN{rng.randrange(10**7, 10**8)}"

            sex_code = rng.choice(NEW_SEX_CODES) if self.chance("new_sex_code") else \
                {"M": "M", "F": "F", "X": "U"}[sex]
            if sex_code in NEW_SEX_CODES:
                self.stats["new_sex_code"] += 1

            new_ehr.append([mrn, given, family, fmt_d(birth), sex_code, hcn,
                            make_postal(rng, fsa), make_phone(rng), "English",
                            "N", self.event_ts()])

            # Most new patients also register in scheduling
            if rng.random() < 0.8:
                new_sched.append([f"SCH{rng.randrange(10**7, 10**8)}", given, family,
                                  birth.strftime("%d/%m/%Y"), "MALE" if sex == "M" else "FEMALE",
                                  f"{hcn[:4]} {hcn[4:7]} {hcn[7:]}",
                                  make_postal(rng, fsa), make_phone(rng), self.event_ts()])
            self.stats["new_patients"] += 1

        # ---- 2. attribute changes on existing patients ---------------
        for r in ehr:
            changed = False
            row = dict(r)

            if self.chance("address_change"):
                row["postal_code"] = make_postal(rng, rng.choice(FSA_POOL))
                changed = True
                self.stats["address_change"] += 1

            if self.chance("surname_change"):
                # Marriage or divorce. Same health card, same birth date,
                # different family name — the MPI must still link them, and
                # it is a genuinely hard case because the name and soundex
                # both change.
                row["last_name"] = rng.choice(FAMILY)
                changed = True
                self.stats["surname_change"] += 1

            if self.chance("language_change"):
                row["preferred_language"] = rng.choice(
                    ["French", "Mandarin", "Punjabi", "Spanish"])
                changed = True

            if changed:
                if self.chance("trailing_whitespace"):
                    row["last_name"] = corrupt_text(rng, row["last_name"])
                    self.stats["text_corruption"] += 1
                row["last_modified_ts"] = self.event_ts()
                new_ehr.append([row.get(c, "") for c in
                                ["patient_id", "first_name", "last_name",
                                 "date_of_birth", "gender", "health_card_number",
                                 "postal_code", "primary_phone", "preferred_language",
                                 "deceased_indicator", "last_modified_ts"]])

        # ---- 3. cross-source appearances -----------------------------
        # An existing patient shows up in a system they were absent from.
        # The MPI should fold this into their existing cluster.
        candidates = [r for r in ehr if r.get("health_card_number")]
        rng.shuffle(candidates)
        for r in candidates[:int(len(candidates) * DELTA_DEFECTS["cross_source_new"])]:
            hcn = r["health_card_number"]
            if rng.random() < 0.5:
                new_sched.append([f"SCH{rng.randrange(10**7, 10**8)}",
                                  NICKNAMES.get(r["first_name"], r["first_name"]),
                                  r["last_name"],
                                  (r["date_of_birth"] or "").replace("-", "/"),
                                  "MALE" if r["gender"] in ("M", "1") else "FEMALE",
                                  hcn, r["postal_code"], r["primary_phone"],
                                  self.event_ts()])
            else:
                new_fin.append([f"ACC{rng.randrange(10**7, 10**8)}",
                                r["first_name"], r["last_name"],
                                (r["date_of_birth"] or ""), r["gender"].lower(),
                                f"{hcn[:4]}-{hcn[4:7]}-{hcn[7:]}" if len(hcn) == 10 else hcn,
                                r["postal_code"], r["phone" ] if "phone" in r else r["primary_phone"],
                                self.event_ts()])
            self.stats["cross_source_new"] += 1

        # ---- 4. new intra-source duplicates --------------------------
        # Re-registered in the EHR under a fresh MRN, with the drift a busy
        # registration desk produces.
        for r in candidates[:int(len(candidates) * DELTA_DEFECTS["new_intra_dup"])]:
            given = NICKNAMES.get(r["first_name"], r["first_name"])
            if rng.random() < 0.4:
                given = typo(rng, given)
            family = r["last_name"]
            if rng.random() < 0.3:
                family = typo(rng, family)
            hcn = r["health_card_number"] if rng.random() < 0.6 else ""
            new_ehr.append([f"MRN{rng.randrange(10**7, 10**8)}", given, family,
                            r["date_of_birth"], r["gender"], hcn,
                            r["postal_code"] if rng.random() < 0.7 else "",
                            r["primary_phone"], "", "N", self.event_ts()])
            self.stats["new_intra_dup"] += 1

        # ---- 5. duplicate delivery -----------------------------------
        # Exact rows re-sent. _row_hash should absorb these without doubling
        # the row count.
        n_dupe = int(len(new_ehr) * DELTA_DEFECTS["duplicate_delivery"])
        if n_dupe:
            new_ehr.extend(rng.sample(new_ehr, min(n_dupe, len(new_ehr))))
            self.stats["duplicate_delivery"] += n_dupe

        self.add("EHR/patient",
                 ["patient_id", "first_name", "last_name", "date_of_birth",
                  "gender", "health_card_number", "postal_code", "primary_phone",
                  "preferred_language", "deceased_indicator", "last_modified_ts"],
                 new_ehr)
        self.add("SCHED/patient",
                 ["patient_ref", "given", "surname", "dob", "sex", "health_card",
                  "postal", "contact_phone", "modified_at"], new_sched)
        self.add("FIN/patient_account",
                 ["account_holder_id", "first_nm", "last_nm", "birth_dt",
                  "gender_cd", "hc_number", "zip_postal", "phone_number",
                  "update_dt"], new_fin)

        self.all_patient_ids = [r[0] for r in new_ehr] + list(ehr_by_id.keys())

    # ================================================== FACILITY / STAFF
    def gen_facility_changes(self):
        """Attribute changes that should create SCD2 versions in Gold."""
        rng = self.rng

        hdr_d, doctors = self.read("HR", "doctor")
        hdr_b, beds = self.read("FACIL", "bed")
        _, depts = self.read("FACIL", "department")
        dept_ids = [d["department_id"] for d in depts]

        new_doctors, new_beds = [], []

        for d in doctors:
            row, changed = dict(d), False
            if self.chance("doctor_dept_move"):
                # A doctor changing department is the canonical SCD2 case.
                # Their 2024 admissions must stay attributed to the old one.
                same_hosp = [x for x in dept_ids if x.startswith(d["hospital_id"])]
                if same_hosp:
                    row["primary_department_id"] = rng.choice(same_hosp)
                    changed = True
                    self.stats["doctor_dept_move"] += 1
            if self.chance("doctor_fte_change"):
                row["fte"] = str(rng.choice([1.0, 0.8, 0.6, 0.5, 0.4]))
                changed = True
                self.stats["doctor_fte_change"] += 1
            if rng.random() < 0.01:
                row["is_active"] = "0"
                changed = True
                self.stats["doctor_deactivated"] += 1
            if changed:
                row["update_ts"] = self.event_ts()
                new_doctors.append([row.get(c, "") for c in hdr_d])

        for b in beds:
            if self.chance("bed_ward_change"):
                row = dict(b)
                row["ward_type"] = rng.choice(
                    ["Med-Surg", "ICU", "Step-Down", "Rehab", "Maternity"])
                row["ward_name"] = f"{row['ward_type'][:4].upper()}-UPD"
                row["update_ts"] = self.event_ts()
                new_beds.append([row.get(c, "") for c in hdr_b])
                self.stats["bed_ward_change"] += 1

        # Facility entities are full snapshots, so emit the whole table with
        # the changes applied — that is what the source system would send.
        changed_docs = {r[0] for r in new_doctors}
        full_doctors = new_doctors + [[d.get(c, "") for c in hdr_d]
                                      for d in doctors if d["doctor_id"] not in changed_docs]
        changed_beds = {r[0] for r in new_beds}
        full_beds = new_beds + [[b.get(c, "") for c in hdr_b]
                                for b in beds if b["bed_id"] not in changed_beds]

        self.add("HR/doctor", hdr_d, full_doctors)
        self.add("FACIL/bed", hdr_b, full_beds)

        for src, ent in [("FACIL", "hospital"), ("FACIL", "department")]:
            h, rows = self.read(src, ent)
            self.add(f"{src}/{ent}", h, [[r.get(c, "") for c in h] for r in rows])

    # ================================================== CLINICAL ACTIVITY
    def gen_activity(self):
        """New encounters for existing and new patients, including
        readmissions that reference discharges from the first batch."""
        rng = self.rng

        _, old_adm = self.read("EHR", "admission")
        _, depts = self.read("FACIL", "department")
        _, beds = self.read("FACIL", "bed")
        _, doctors = self.read("HR", "doctor")

        inpatient = [d for d in depts if d.get("is_inpatient_unit") == "1"]
        ed_depts = [d for d in depts if d.get("department_name") == "Emergency"]
        beds_by_dept = defaultdict(list)
        for b in beds:
            beds_by_dept[b["department_id"]].append(b)
        docs_by_dept = defaultdict(list)
        for d in doctors:
            docs_by_dept[d["primary_department_id"]].append(d)

        # Highest existing ids, so the delta continues the sequence rather
        # than colliding.
        def max_seq(rows, col, prefix):
            vals = [int(r[col][len(prefix):]) for r in rows
                    if r.get(col, "").startswith(prefix) and r[col][len(prefix):].isdigit()]
            return max(vals) if vals else 0

        seq = {
            "adm": max_seq(old_adm, "admission_id", "ADM"),
            "enc": 0, "ed": 0, "lab": 0, "med": 0, "clm": 0, "inv": 0,
            "srv": 0, "appt": 0,
        }
        _, old_ed = self.read("EHR", "emergency_visit")
        seq["ed"] = max_seq(old_ed, "ed_visit_id", "EDV")
        seq["enc"] = max_seq(old_ed, "encounter_id", "ENC")
        _, old_lab = self.read("LIS", "lab_result")
        seq["lab"] = max_seq(old_lab, "lab_result_id", "LAB")
        _, old_med = self.read("PHARM", "medication_order")
        seq["med"] = max_seq(old_med, "medication_order_id", "MED")
        _, old_appt = self.read("SCHED", "appointment")
        seq["appt"] = max_seq(old_appt, "appointment_id", "APT")
        _, old_inv = self.read("FIN", "invoice")
        seq["inv"] = max_seq(old_inv, "invoice_id", "INV")
        _, old_srv = self.read("SURVEY", "survey_response")
        seq["srv"] = max_seq(old_srv, "survey_response_id", "SRV")

        def nid(k, p, w=8):
            seq[k] += 1
            return f"{p}{seq[k]:0{w}d}"

        # Patients discharged near the end of the first batch are the
        # readmission candidates — this is what links the two batches.
        recent_discharges = []
        for r in old_adm:
            if r.get("discharge_ts"):
                try:
                    d = datetime.strptime(r["discharge_ts"][:10], "%Y-%m-%d").date()
                    if (self.window_start - d).days <= 45:
                        recent_discharges.append((r["patient_id"], d, r))
                except ValueError:
                    pass

        _, ehr_pat = self.read("EHR", "patient")
        patient_ids = [r["patient_id"] for r in ehr_pat]
        new_patient_ids = [r[0] for r in self.tables["EHR/patient"][1]]
        active = list(set(patient_ids + new_patient_ids))

        adm_rows, ed_rows, bed_rows, dx_rows = [], [], [], []
        lab_rows, med_rows, appt_rows, appt_hist = [], [], [], []
        inv_rows, inv_line_rows, srv_rows = [], [], []
        payer_ids = [p[0] for p in PAYERS]
        payer_w = [p[4] for p in PAYERS]

        # ---- readmissions tied to the previous batch ------------------
        n_readmit = int(len(recent_discharges) * 0.14)
        for pid, disc_date, prior in rng.sample(
                recent_discharges, min(n_readmit, len(recent_discharges))):
            gap = rng.randrange(1, 30)
            adm_dt = datetime.combine(disc_date + timedelta(days=gap),
                                      datetime.min.time()) + timedelta(
                                          hours=rng.randrange(24))
            dept = rng.choice(inpatient) if inpatient else None
            if not dept:
                continue
            los = lognormal_days(rng, 1.7, 0.7)
            disc = adm_dt + timedelta(days=los)
            adm_id = nid("adm", "ADM")
            enc_id = nid("enc", "ENC")
            doc = rng.choice(docs_by_dept.get(dept["department_id"]) or doctors)
            dx = rng.choice(DIAGNOSES)
            adm_rows.append([adm_id, enc_id, pid, doc["doctor_id"],
                             dept["department_id"], dept["hospital_id"],
                             fmt_ts(adm_dt), fmt_ts(disc),
                             fmt_ts(adm_dt + timedelta(days=los)),
                             "E", "HOME", dx[0],
                             rng.choices(payer_ids, weights=payer_w)[0],
                             fmt_ts(disc + timedelta(hours=3))])
            dx_rows.append([enc_id, pid, dx[0], 1, "Discharge", 1, 1,
                            fmt_d(adm_dt.date())])
            self.stats["readmissions"] += 1

        # ---- ordinary new admissions ----------------------------------
        # Readmissions alone would leave fact_admission almost empty and
        # would never exercise bed assignments, billing or surveys.
        n_adm = max(100, int(len(active) * 0.06))
        for _ in range(n_adm):
            pid = rng.choice(active)
            dept = rng.choice(inpatient) if inpatient else None
            if not dept:
                break
            adm_dt = datetime.strptime(self.event_ts(), "%Y-%m-%d %H:%M:%S")
            los = lognormal_days(rng, 1.6, 0.7)
            still_open = rng.random() < 0.08
            disc = None if still_open else adm_dt + timedelta(
                days=los, hours=rng.randrange(0, 12))
            adm_id = nid("adm", "ADM")
            enc_id = nid("enc", "ENC")
            doc = rng.choice(docs_by_dept.get(dept["department_id"]) or doctors)
            dx = (rng.choice(ORPHAN_DIAGNOSES)
                  if self.chance("unknown_diagnosis") else rng.choice(DIAGNOSES)[0])
            if dx in ORPHAN_DIAGNOSES:
                self.stats["unknown_diagnosis"] += 1
            payer = rng.choices(payer_ids, weights=payer_w)[0]
            disp = "" if still_open else rng.choices(
                ["HOME", "HMCR", "REHB", "LTC", "TRAN", "EXP"],
                weights=[72, 12, 6, 5, 3, 2])[0]

            adm_rows.append([adm_id, enc_id, pid, doc["doctor_id"],
                             dept["department_id"], dept["hospital_id"],
                             fmt_ts(adm_dt), fmt_ts(disc) if disc else "",
                             fmt_ts(adm_dt + timedelta(days=los)),
                             rng.choices(["E", "U", "EL"], weights=[60, 20, 20])[0],
                             disp, dx, payer,
                             fmt_ts((disc or adm_dt) + timedelta(hours=3))])

            dx_rows.append([enc_id, pid, dx, 1, "Discharge", 1, 1,
                            fmt_d(adm_dt.date())])
            for rank, cd in enumerate(rng.sample(
                    [d for d in DIAGNOSES if d[4] == 1],
                    min(rng.randrange(0, 4), 10)), 2):
                dx_rows.append([enc_id, pid, cd[0], rank, "Comorbid", 0, 1,
                                fmt_d(adm_dt.date())])

            # Bed assignments — one per unit, more when transferred
            pool = beds_by_dept.get(dept["department_id"], [])
            if pool and disc:
                n_moves = rng.choices([1, 2, 3], weights=[76, 19, 5])[0]
                cursor = adm_dt
                total_s = (disc - adm_dt).total_seconds()
                for m in range(n_moves):
                    bed = rng.choice(pool)
                    end = disc if m == n_moves - 1 else cursor + timedelta(
                        seconds=total_s * rng.uniform(0.2, 0.7) / n_moves)
                    bed_rows.append([adm_id, enc_id, bed["bed_id"],
                                     fmt_ts(cursor), fmt_ts(end),
                                     "Transfer" if m else "Admission"])
                    cursor = end

            # Invoice and lines
            inv_id = nid("inv", "INV")
            net_total = 0.0
            for ln in range(1, rng.randrange(2, 8)):
                cat, lo, hi = rng.choice([
                    ("Room and board", 850, 2400), ("Surgical procedure", 1800, 14000),
                    ("Diagnostic imaging", 220, 1900), ("Laboratory", 35, 480),
                    ("Pharmacy", 25, 2200), ("Physician services", 180, 1400)])
                charge = round(rng.uniform(lo, hi) * (1 + los * 0.08), 2)
                discount = round(charge * rng.choice([0, 0, 0.05, 0.1]), 2)
                net = round(charge - discount, 2)
                payment = round(net * rng.choice([0, 0.4, 0.75, 1.0]), 2)
                net_total += net
                inv_line_rows.append([inv_id, ln, pid, dept["department_id"],
                                      dept["hospital_id"], payer, enc_id,
                                      f"SC{rng.randrange(1000, 9999)}", cat,
                                      rng.choice([1, 1, 2, 3]), charge, discount,
                                      0.0, net, payment,
                                      round(max(0, net - payment), 2)])
            inv_rows.append([inv_id, pid, enc_id, dept["hospital_id"], payer,
                             fmt_d((disc or adm_dt).date()), fmt_d(adm_dt.date()),
                             round(net_total, 2),
                             fmt_ts((disc or adm_dt) + timedelta(days=1))])

            # Satisfaction survey
            if disc and rng.random() < 0.4:
                base_score = 8.4 - min(1.5, los * 0.06)
                def sc(off=0.0, _b=base_score):
                    return max(1, min(10, int(round(rng.gauss(_b + off, 1.35)))))
                overall = sc()
                nps = ("Promoter" if overall >= 9 else
                       "Passive" if overall >= 7 else "Detractor")
                srv_rows.append([nid("srv", "SRV"), pid, doc["doctor_id"],
                                 dept["department_id"], dept["hospital_id"], enc_id,
                                 fmt_d((disc + timedelta(days=rng.randrange(2, 21))).date()),
                                 fmt_d(disc.date()), "Inpatient",
                                 overall, sc(-0.9), sc(0.6), sc(0.3), sc(0.1),
                                 sc(-0.3), sc(0.2), nps,
                                 1 if rng.random() < 0.42 else 0])
            self.stats["new_admissions"] += 1

        # ---- ordinary new activity ------------------------------------
        n_encounters = max(200, int(len(active) * 0.10))
        for _ in range(n_encounters):
            pid = rng.choice(active)
            ts = self.event_ts()
            base_dt = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")

            # ED visit
            if ed_depts and rng.random() < 0.55:
                dept = rng.choice(ed_depts)
                ctas = rng.choices([1, 2, 3, 4, 5],
                                   weights=[1.2, 14.8, 45.2, 29.8, 9.0])[0]
                arrival = base_dt.replace(hour=diurnal_hour(rng))
                triage = arrival + timedelta(minutes=rng.randrange(2, 40))
                seen = triage + timedelta(minutes=rng.randrange(10, 300))
                depart = seen + timedelta(minutes=rng.randrange(30, 400))
                dx = (rng.choice(ORPHAN_DIAGNOSES)
                      if self.chance("unknown_diagnosis") else rng.choice(DIAGNOSES)[0])
                if dx in ORPHAN_DIAGNOSES:
                    self.stats["unknown_diagnosis"] += 1
                ed_rows.append([nid("ed", "EDV"), nid("enc", "ENC"), pid,
                                dept["hospital_id"], dept["department_id"],
                                fmt_ts(arrival), fmt_ts(triage), fmt_ts(seen),
                                "", "", fmt_ts(depart), ctas,
                                rng.choice(["Walk-in", "Ambulance", "Transfer"]),
                                dx, 0, 0,
                                rng.choices(payer_ids, weights=payer_w)[0],
                                fmt_ts(depart + timedelta(hours=2))])

            # Lab results
            for _ in range(rng.randrange(0, 4)):
                test = rng.choice(LAB_TESTS)
                loinc = (rng.choice(ORPHAN_LOINC)
                         if self.chance("unknown_loinc") else test[0])
                if loinc in ORPHAN_LOINC:
                    self.stats["unknown_loinc"] += 1
                order = base_dt + timedelta(hours=rng.randrange(0, 48))
                collect = order + timedelta(minutes=rng.randrange(5, 90))
                result = collect + timedelta(minutes=rng.randrange(20, 300))
                val = round(rng.uniform(0.5, 200), 3) if test[5] is not None else ""
                dept = rng.choice(inpatient) if inpatient else None
                lab_rows.append([nid("lab", "LAB"), "", pid,
                                 rng.choice(doctors)["doctor_id"],
                                 dept["department_id"] if dept else "",
                                 dept["hospital_id"] if dept else "",
                                 loinc, fmt_ts(order), fmt_ts(collect),
                                 fmt_ts(result),
                                 rng.choices(["Routine", "STAT"], weights=[78, 22])[0],
                                 val, "", rng.choice(["N", "H", "L", "HH"]),
                                 0, fmt_ts(result + timedelta(minutes=5))])

            # Medication orders
            for _ in range(rng.randrange(0, 3)):
                med = rng.choice(MEDICATIONS)
                din = rng.choice(ORPHAN_DIN) if self.chance("unknown_din") else med[0]
                if din in ORPHAN_DIN:
                    self.stats["unknown_din"] += 1
                order = base_dt + timedelta(hours=rng.randrange(0, 48))
                disp = order + timedelta(minutes=rng.randrange(5, 200))
                qty = rng.choice([1, 2, 5, 10, 30])
                cost = round(rng.uniform(0.5, 80), 2)
                total = f"{qty * cost:,.2f}" if self.chance("thousand_separator") \
                    else f"{qty * cost:.2f}"
                if "," in total:
                    self.stats["thousand_separator"] += 1
                dept = rng.choice(inpatient) if inpatient else None
                med_rows.append([nid("med", "MED"), "", pid,
                                 rng.choice(doctors)["doctor_id"],
                                 dept["department_id"] if dept else "",
                                 dept["hospital_id"] if dept else "",
                                 din, fmt_ts(order), fmt_ts(disp),
                                 rng.choice([5, 10, 25, 50, 100]),
                                 rng.choice(["mg", "mL", "units"]),
                                 rng.choice(["OD", "BID", "TID", "PRN"]),
                                 qty, qty, rng.choice([1, 7, 30]),
                                 cost, total, 0,
                                 fmt_ts(disp + timedelta(minutes=10))])

            # Appointments
            for _ in range(rng.randrange(0, 3)):
                dept = rng.choice([d for d in depts
                                   if d.get("is_clinical") == "1"] or depts)
                doc = rng.choice(docs_by_dept.get(dept["department_id"]) or doctors)
                lead = rng.randrange(1, 60)
                sched_dt = base_dt + timedelta(days=rng.randrange(0, 40))
                hour, minute = business_hour(rng)
                sched_dt = sched_dt.replace(hour=hour, minute=minute)
                status = rng.choices(
                    ["COMPLETED", "NO_SHOW", "CANCELLED_PATIENT", "SCHEDULED"],
                    weights=[70, 9, 12, 9])[0]
                appt_id = nid("appt", "APT")
                appt_rows.append([appt_id, pid, doc["doctor_id"],
                                  dept["department_id"], dept["hospital_id"],
                                  fmt_ts(sched_dt - timedelta(days=lead)),
                                  fmt_ts(sched_dt),
                                  rng.choice([15, 20, 30, 45]),
                                  rng.choice(["Follow-up", "New consultation",
                                              "Procedure"]),
                                  0, rng.choice([0, 1]),
                                  fmt_ts(sched_dt) if status == "COMPLETED" else "",
                                  fmt_ts(sched_dt) if status == "COMPLETED" else "",
                                  fmt_ts(sched_dt + timedelta(minutes=30))
                                  if status == "COMPLETED" else "",
                                  "", "", status,
                                  fmt_ts(sched_dt + timedelta(days=1))])
                appt_hist.append([appt_id, status, fmt_ts(sched_dt)])

        self.add("EHR/admission",
                 ["admission_id", "encounter_id", "patient_id",
                  "attending_physician_id", "department_id", "hospital_id",
                  "admission_ts", "discharge_ts", "expected_discharge_ts",
                  "admission_type", "discharge_disposition_code",
                  "primary_diagnosis_code", "payer_id", "last_modified_ts"],
                 adm_rows)
        self.add("EHR/emergency_visit",
                 ["ed_visit_id", "encounter_id", "patient_id", "hospital_id",
                  "department_id", "arrival_ts", "triage_ts", "physician_seen_ts",
                  "admit_decision_ts", "bed_assigned_ts", "departure_ts",
                  "triage_score", "arrival_mode", "triage_diagnosis_code",
                  "left_without_being_seen", "resulted_in_admission", "payer_id",
                  "last_modified_ts"], ed_rows)
        self.add("EHR/diagnosis",
                 ["encounter_id", "patient_id", "diagnosis_code", "diagnosis_rank",
                  "diagnosis_type", "is_primary", "is_present_on_admission",
                  "diagnosis_date"], dx_rows)
        self.add("EHR/bed_assignment",
                 ["admission_id", "encounter_id", "bed_id", "assignment_start_ts",
                  "assignment_end_ts", "assignment_reason"], bed_rows)
        self.add("LIS/lab_result",
                 ["lab_result_id", "encounter_id", "patient_id",
                  "ordering_doctor_id", "department_id", "hospital_id",
                  "loinc_code", "order_ts", "collect_ts", "result_ts", "priority",
                  "result_value_numeric", "result_value_text", "abnormal_flag",
                  "is_critical", "modified_at"], lab_rows)
        self.add("PHARM/medication_order",
                 ["medication_order_id", "encounter_id", "patient_id",
                  "prescribing_doctor_id", "department_id", "hospital_id", "din",
                  "order_ts", "dispense_ts", "dose_amount", "dose_unit",
                  "frequency_code", "quantity_ordered", "quantity_dispensed",
                  "days_supply", "unit_cost", "total_cost", "is_discontinued",
                  "updated_at"], med_rows)
        self.add("SCHED/appointment",
                 ["appointment_id", "patient_ref", "doctor_id", "department_id",
                  "hospital_id", "booking_ts", "scheduled_ts",
                  "scheduled_duration_min", "appointment_type", "is_first_visit",
                  "is_virtual", "checkin_ts", "seen_ts", "checkout_ts",
                  "cancellation_ts", "cancelled_by", "status_code", "modified_at"],
                 appt_rows)
        self.add("SCHED/appointment_status_history",
                 ["appointment_id", "status_code", "status_ts"], appt_hist)
        self.add("FIN/invoice",
                 ["invoice_id", "patient_id", "encounter_id", "hospital_id",
                  "payer_id", "invoice_date", "service_date", "total_amount",
                  "update_dt"], inv_rows)
        self.add("FIN/invoice_line",
                 ["invoice_id", "invoice_line_number", "patient_id",
                  "department_id", "hospital_id", "payer_id", "encounter_id",
                  "service_code", "service_category", "quantity", "charge_amount",
                  "discount_amount", "tax_amount", "net_amount", "payment_amount",
                  "outstanding_amount"], inv_line_rows)
        self.add("SURVEY/survey_response",
                 ["survey_response_id", "patient_id", "doctor_id", "department_id",
                  "hospital_id", "encounter_id", "response_date", "service_date",
                  "encounter_type", "overall_score", "wait_time_score",
                  "staff_courtesy_score", "cleanliness_score",
                  "communication_score", "pain_management_score",
                  "would_recommend_score", "nps_category",
                  "has_free_text_comment"], srv_rows)

    # ================================================== OUTPUT
    def write(self, ingest_date: str):
        os.makedirs(self.out, exist_ok=True)
        manifest = []
        for name, (header, rows) in sorted(self.tables.items()):
            if not header:
                continue
            source, entity = name.split("/")
            folder = os.path.join(self.out, source, entity,
                                  f"ingest_date={ingest_date}")
            os.makedirs(folder, exist_ok=True)
            path = os.path.join(folder, f"{entity}_{ingest_date}.csv")
            with open(path, "w", newline="", encoding="utf-8") as fh:
                w = csv.writer(fh)
                w.writerow(header)
                w.writerows(rows)
            manifest.append((source, entity, len(rows)))
        return manifest


# =====================================================================
# APPLY TO THE FIVE SOURCE SYSTEMS
# =====================================================================

def apply_to_sources(batch: DeltaBatch, sources_dir: str, ingest_date: str):
    """Distribute the delta to where each pipeline will read it.

    SQL Server gets INSERT scripts rather than a bulk reload, because the
    point is to test incremental extraction — a truncate-and-reload would
    make every watermark meaningless.
    """
    print("\nDistributing to source systems")

    # ---- SQL Server: INSERT script ----------------------------------
    sql_dir = os.path.join(sources_dir, "01-sqlserver")
    os.makedirs(sql_dir, exist_ok=True)
    db_map = {
        "EHR": ("OHN_EHR", ["patient", "admission", "bed_assignment",
                            "diagnosis", "emergency_visit"]),
        "SCHED": ("OHN_SCHED", ["patient", "appointment",
                                "appointment_status_history"]),
        "FIN": ("OHN_FIN", ["patient_account", "invoice", "invoice_line"]),
    }
    lines = ["/* Delta batch — INSERT only, so watermarks stay meaningful. */",
             "SET NOCOUNT ON;", "GO", ""]
    n_sql = 0
    for src, (db, entities) in db_map.items():
        lines += [f"USE {db};", "GO", ""]
        for ent in entities:
            key = f"{src}/{ent}"
            if key not in batch.tables:
                continue
            header, rows = batch.tables[key]
            if not rows:
                continue
            cols = ", ".join(f"[{c}]" for c in header)
            for chunk_start in range(0, len(rows), 500):
                chunk = rows[chunk_start:chunk_start + 500]
                values = []
                for r in chunk:
                    vals = ", ".join(
                        "NULL" if v == "" or v is None
                        else "'" + str(v).replace("'", "''") + "'" for v in r)
                    values.append(f"({vals})")
                lines.append(f"INSERT INTO dbo.{ent} ({cols}) VALUES")
                lines.append(",\n".join(values) + ";")
                lines.append("GO")
                n_sql += len(chunk)
            lines.append("")
    with open(os.path.join(sql_dir, "05_delta_batch.sql"), "w",
              encoding="utf-8") as fh:
        fh.write("\n".join(lines))
    print(f"  SQL Server   05_delta_batch.sql        {n_sql:>8,} rows")

    # ---- REST API: append to the JSON payloads ----------------------
    api_data = os.path.join(sources_dir, "02-api", "data")
    n_api = 0
    for key, fname in [("PHARM/medication_order", "pharm_medication_order.json"),
                       ("FACIL/hospital", "facil_hospital.json"),
                       ("FACIL/department", "facil_department.json"),
                       ("FACIL/bed", "facil_bed.json")]:
        if key not in batch.tables:
            continue
        header, rows = batch.tables[key]
        if not rows:
            continue
        path = os.path.join(api_data, fname)
        existing = []
        if os.path.exists(path):
            with open(path, encoding="utf-8") as fh:
                existing = json.load(fh)
        new = [{h: (v if v != "" else None) for h, v in zip(header, r)} for r in rows]
        # Facility entities are full snapshots — replace. Medication orders
        # are incremental — append, keyed on their id column so re-applying
        # the same delta (e.g. after an earlier partial/failed run) upserts
        # instead of duplicating rows.
        if key.startswith("FACIL"):
            merged = new
        else:
            id_col = header[0]
            by_id = {r[id_col]: r for r in existing}
            by_id.update({r[id_col]: r for r in new})
            merged = list(by_id.values())
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(merged, fh)
        n_api += len(new)
        print(f"  API          {fname:<28}{len(new):>8,} rows")
    if n_api:
        print("  -> restart api_server.py to serve the updated payloads")

    # ---- Excel: rewrite the workbooks -------------------------------
    try:
        from openpyxl import load_workbook
        sp_dir = os.path.join(sources_dir, "05-sharepoint")
        for key, fname in [("HR/doctor", "doctor_roster.xlsx")]:
            if key not in batch.tables or not os.path.exists(
                    os.path.join(sp_dir, fname)):
                continue
            header, rows = batch.tables[key]
            wb = load_workbook(os.path.join(sp_dir, fname))
            ws = wb.active
            ws.delete_rows(5, ws.max_row)
            for r_ix, row in enumerate(rows, 5):
                for c_ix, val in enumerate(row, 1):
                    ws.cell(row=r_ix, column=c_ix, value=val)
            wb.save(os.path.join(sp_dir, fname))
            print(f"  Excel        {fname:<28}{len(rows):>8,} rows")
    except ImportError:
        print("  Excel        openpyxl not installed — skipped")

    print("\n  Reference CSVs unchanged — orphan codes are deliberately absent,")
    print("  so they resolve to the unknown member in Gold.")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--landing", default="./landing",
                    help="the ORIGINAL landing folder, read for existing ids")
    ap.add_argument("--out", default="./landing-delta")
    ap.add_argument("--days", type=int, default=60,
                    help="width of the delta window in days")
    ap.add_argument("--ingest-date",
                    default=date.today().strftime("%Y-%m-%d"))
    ap.add_argument("--apply", default=None,
                    help="path to the sources/ folder; also updates the five systems")
    ap.add_argument("--seed", type=int, default=SEED)
    ap.add_argument("--clean", action="store_true")
    args = ap.parse_args()

    if args.clean and os.path.isdir(args.out):
        shutil.rmtree(args.out)

    print(f"Delta batch — {args.days}-day window ending {args.ingest_date}")
    print(f"Reading existing ids from {args.landing}\n")

    b = DeltaBatch(args.landing, args.out, args.days, args.seed)
    print("  patients and identity changes...")
    b.gen_patients()
    print("  facility and staff changes...")
    b.gen_facility_changes()
    print("  clinical activity...")
    b.gen_activity()
    manifest = b.write(args.ingest_date)

    print(f"\n{'Source':<10}{'Entity':<30}{'Rows':>10}")
    print("-" * 52)
    total = 0
    for src, ent, n in manifest:
        print(f"{src:<10}{ent:<30}{n:>10,}")
        total += n
    print("-" * 52)
    print(f"{'TOTAL':<40}{total:>10,}")

    print("\nInconsistencies introduced:")
    for k in sorted(b.stats):
        print(f"  {k:<26}{b.stats[k]:>8,}")

    if args.apply:
        apply_to_sources(b, args.apply, args.ingest_date)

    print(f"\nWritten to {os.path.abspath(args.out)}")
    print("\nWhat to check after running the pipelines — see")
    print("docs/11-delta-batch-testing.md")


if __name__ == "__main__":
    main()